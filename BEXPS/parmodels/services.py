import hashlib
import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode

import requests
from django.conf import settings
from django.core.cache import cache
from django.core.mail import send_mail
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import OrganizationMembership

CURRENT_ORGANIZATION_SESSION_KEY = "current_organization_id"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CACHE_MISSING = object()
logger = logging.getLogger(__name__)


def build_cache_key(namespace, organization_id, *parts, params=None):
    raw_parts = [str(namespace), f"org:{organization_id}"]
    raw_parts.extend(str(part) for part in parts if part not in (None, ""))

    if params is not None:
        if hasattr(params, "lists"):
            query_parts = [
                (key, value)
                for key, values in params.lists()
                for value in values
            ]
        else:
            query_parts = list(params.items())
        raw_parts.append(urlencode(sorted(query_parts), doseq=True))

    raw_value = "|".join(raw_parts)
    digest = hashlib.sha256(raw_value.encode("utf-8")).hexdigest()[:20]
    return f"{namespace}:org:{organization_id}:{digest}"


def cache_get_or_set(key, factory, timeout):
    value = cache.get(key, _CACHE_MISSING)
    if value is not _CACHE_MISSING:
        return value
    value = factory()
    cache.set(key, value, timeout)
    return value


def _format_xlsx_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, Decimal):
        return float(value)
    return value


def build_xlsx_response(filename, sheet_title, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = (sheet_title or "Export")[:31]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    column_widths = [len(str(header)) for header in headers]
    for row in rows:
        formatted_row = [_format_xlsx_value(value) for value in row]
        worksheet.append(formatted_row)
        for index, value in enumerate(formatted_row):
            column_widths[index] = max(column_widths[index], len(str(value)))

    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 45)

    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def geocode_yandex_address(address):
    address = (address or "").strip()
    if not address or not settings.YANDEX_GEOCODER_API_KEY:
        return None

    try:
        response = requests.get(
            "https://geocode-maps.yandex.ru/1.x/",
            params={
                "apikey": settings.YANDEX_GEOCODER_API_KEY,
                "geocode": address,
                "format": "json",
                "results": 1,
            },
            timeout=settings.YANDEX_GEOCODER_TIMEOUT,
        )
        response.raise_for_status()
        data = response.json()
        feature_members = (
            data
            .get("response", {})
            .get("GeoObjectCollection", {})
            .get("featureMember", [])
        )
        if not feature_members:
            return None

        pos = (
            feature_members[0]
            .get("GeoObject", {})
            .get("Point", {})
            .get("pos", "")
        )
        longitude_raw, latitude_raw = pos.split()
        latitude = Decimal(latitude_raw).quantize(Decimal("0.000001"))
        longitude = Decimal(longitude_raw).quantize(Decimal("0.000001"))
        return latitude, longitude
    except (requests.RequestException, ValueError, KeyError, InvalidOperation) as exc:
        logger.warning("Yandex geocoding failed for address '%s': %s", address, exc)
        return None


def get_current_membership(user, request=None):
    if not user.is_authenticated:
        return None

    memberships = (
        OrganizationMembership.objects
        .filter(user=user, status=OrganizationMembership.Status.ACTIVE)
        .select_related("organization", "department")
        .order_by("id")
    )

    if request is not None:
        organization_id = request.session.get(CURRENT_ORGANIZATION_SESSION_KEY)
        if organization_id:
            membership = memberships.filter(organization_id=organization_id).first()
            if membership:
                return membership
            request.session.pop(CURRENT_ORGANIZATION_SESSION_KEY, None)

    return memberships.first()


def get_current_organization(user, request=None):
    membership = get_current_membership(user, request=request)
    if membership:
        return membership.organization
    return None


def user_is_org_admin(user, request=None):
    membership = get_current_membership(user, request=request)
    return bool(
        membership
        and membership.role == OrganizationMembership.Role.ADMIN
    )


def send_organization_invitation_email(invitation, invite_url):
    inviter = invitation.invited_by
    invited_by = ""
    if inviter:
        invited_by = inviter.get_full_name() or inviter.email or inviter.username

    message = render_to_string(
        "emails/organization_invitation.txt",
        {
            "invitation": invitation,
            "organization": invitation.organization,
            "role": invitation.get_role_display(),
            "invite_url": invite_url,
            "expires_at": invitation.expires_at,
            "invited_by": invited_by,
        },
    )
    return send_mail(
        subject="Приглашение в организацию",
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[invitation.email],
        fail_silently=False,
    )
